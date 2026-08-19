"""TSS standards mapping for CAD drawings.

Identifies which TSS standards apply to a drawing using two-step engineering
deductive reasoning (not literal keyword matching):

  Step 1 — Composition Inference: determine the physical nature, material
            and function of the component within the compressor ecosystem.
  Step 2 — Category Mapping: associate the component to the correct
            Applicability category in the TSS standards table.

Entry point for the main pipeline:
    from src.standards.tss_mapper import run_tss_mapping, StandardsMappingResult

Deterministic helpers (no LLM):
    filter_generic_standards(norms)   — removes generic entries like "ISO STANDARDS"
    compare_detected_vs_suggested(detected, mapping_result) → StandardsComparisonResult
"""

from __future__ import annotations

import base64
import logging
import os
import time
from dataclasses import dataclass, field
from pathlib import Path

from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

GCP_PROJECT = os.getenv("GCP_PROJECT_ID", "acim-global-data-lake-sandbox")
GCP_REGION  = os.getenv("GCP_REGION", "us-east5")
MODEL_ID    = "gemini-2.5-flash"

DEFAULT_NORMAS_PATH = Path("normas.xlsx")

# Google Sheets configuration (optional, fallback to local xlsx)
SHEETS_CREDENTIALS_PATH = os.getenv("GOOGLE_SHEETS_CREDENTIALS_PATH")
SHEETS_SPREADSHEET_ID   = os.getenv("GOOGLE_SHEETS_SPREADSHEET_ID")
SHEETS_RANGE            = os.getenv("GOOGLE_SHEETS_RANGE", "Notes!B2:F")

# Confidence threshold: below this the fallback (text + image) is triggered.
CONFIDENCE_THRESHOLD = 0.70

# Generic norm entries that should be filtered out deterministically.
# Patterns are matched case-insensitively against the full norm string.
_GENERIC_NORM_PATTERNS: tuple[str, ...] = (
    "iso standards",
    "iso standard",
    "see iso",
)


# ---------------------------------------------------------------------------
# Standards table loader
# ---------------------------------------------------------------------------

def _load_from_sheets(
    credentials_path: str,
    spreadsheet_id: str,
    range_name: str,
) -> list[dict]:
    """Load standards table from Google Sheets using a service account.

    Args:
        credentials_path: Path to service account JSON key file.
        spreadsheet_id:   The spreadsheet ID (from the URL).
        range_name:       Range to read (e.g., "Notes!B2:F").

    Returns:
        List of dicts with keys: standard, content, category, compressor_series, applicability.
    """
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    logger.info("Loading standards from Google Sheets: %s (range: %s)", spreadsheet_id, range_name)

    # Authenticate with service account
    creds = service_account.Credentials.from_service_account_file(
        credentials_path,
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
    )

    # Build the Sheets API client
    service = build("sheets", "v4", credentials=creds, cache_discovery=False)

    # Read the range
    result = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=range_name,
    ).execute()

    rows = result.get("values", [])
    if not rows:
        logger.warning("No data found in Sheets range: %s", range_name)
        return []

    # Parse rows (skip header if present)
    standards: list[dict] = []
    for i, row in enumerate(rows):
        # Skip header row (if B2 is the first data row, row[0] should be a TSS code)
        if i == 0 and row and "standard" in str(row[0]).lower():
            continue

        if len(row) < 5:
            # Pad missing columns
            row.extend([""] * (5 - len(row)))

        standard, content, category, compressor_series, applicability = row[:5]

        if not standard or not str(standard).strip():
            continue

        standards.append({
            "standard":          str(standard).strip(),
            "content":           str(content).strip()           if content           else "",
            "category":          str(category).strip()          if category          else "",
            "compressor_series": str(compressor_series).strip() if compressor_series else "All",
            "applicability":     str(applicability).strip()     if applicability     else "",
        })

    logger.info("Loaded %d standards from Google Sheets", len(standards))
    return standards


def load_standards_table(xlsx_path: Path = DEFAULT_NORMAS_PATH) -> list[dict]:
    """Load the TSS standards table from Google Sheets or local xlsx.

    Priority:
        1. Google Sheets (if GOOGLE_SHEETS_CREDENTIALS_PATH and GOOGLE_SHEETS_SPREADSHEET_ID are set)
        2. Local xlsx file (fallback)

    Expected sheet layout (sheet "Notes"):
        Row 1: empty
        Row 2: headers  (col B=Standard, C=Content, D=Category,
                          E=Compressor Series, F=Applicability)
        Row 3+: data rows

    Returns:
        List of dicts with keys: standard, content, category, compressor_series, applicability.
    """
    # Try Google Sheets first if configured
    if SHEETS_CREDENTIALS_PATH and SHEETS_SPREADSHEET_ID:
        try:
            return _load_from_sheets(
                SHEETS_CREDENTIALS_PATH,
                SHEETS_SPREADSHEET_ID,
                SHEETS_RANGE,
            )
        except Exception as e:
            logger.warning(
                "Failed to load from Google Sheets (falling back to local xlsx): %s",
                e,
                exc_info=True,
            )

    # Fallback to local xlsx
    import openpyxl

    if not xlsx_path.exists():
        raise FileNotFoundError(
            f"Standards file not found: {xlsx_path}\n"
            "Check the path or use the normas_path parameter."
        )

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Notes"] if "Notes" in wb.sheetnames else wb.active

    standards: list[dict] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        _, standard, content, category, compressor_series, applicability = row[:6]
        if not standard:
            continue
        standards.append({
            "standard":          str(standard).strip(),
            "content":           str(content).strip()           if content           else "",
            "category":          str(category).strip()          if category          else "",
            "compressor_series": str(compressor_series).strip() if compressor_series else "All",
            "applicability":     str(applicability).strip()     if applicability     else "",
        })

    wb.close()
    logger.info("Standards table loaded: %d entries from '%s'", len(standards), xlsx_path)
    return standards


# ---------------------------------------------------------------------------
# Deterministic filter — generic norm strings
# ---------------------------------------------------------------------------

def filter_generic_standards(norms: list[str]) -> tuple[list[str], list[str]]:
    """Remove generic/non-specific norm entries from a list.

    Returns:
        (kept, filtered_out)  — both are lists of strings.

    Examples of filtered entries: "ISO STANDARDS", "SEE ISO STANDARDS",
    "ISO STANDARD", "ISO STANDARDS, SEE TSS 002513".
    """
    kept: list[str] = []
    filtered_out: list[str] = []
    for norm in norms:
        lower = norm.lower().strip()
        is_generic = any(pattern in lower for pattern in _GENERIC_NORM_PATTERNS)
        if is_generic:
            filtered_out.append(norm)
        else:
            kept.append(norm)
    return kept, filtered_out


# ---------------------------------------------------------------------------
# Pydantic models — LLM structured output
# ---------------------------------------------------------------------------

class AppliedStandard(BaseModel):
    """A single TSS standard identified as applicable to the drawing."""

    standard: str = Field(description="Standard code, e.g. TSS 002369")
    applicability_match: str = Field(
        description="Excerpt from the Applicability column that justifies the match"
    )
    reasoning: str = Field(
        description=(
            "Two-step engineering reasoning: "
            "(1) physical nature of the component, "
            "(2) why this standard applies"
        )
    )
    confidence: float = Field(description="Confidence 0.0–1.0 for this specific match")


class StandardsMappingOutput(BaseModel):
    """Structured LLM output for TSS standards mapping."""

    component_title: str = Field(
        description="Title extracted from the TITLE, DOCUMENT TYPE field of the drawing block"
    )
    component_inference: str = Field(
        description=(
            "Composition inference: physical nature, material and function "
            "of the component within the compressor ecosystem"
        )
    )
    applied_standards: list[AppliedStandard] = Field(
        description="TSS standards identified as applicable, in order of relevance"
    )
    explicitly_excluded: list[str] = Field(
        description="Standards considered and explicitly discarded, with brief justification"
    )
    overall_confidence: float = Field(
        description="Overall confidence 0.0–1.0 for the complete mapping"
    )
    needs_visual_confirmation: bool = Field(
        description="True if the vectorised text was insufficient and image analysis would help"
    )


# ---------------------------------------------------------------------------
# Result dataclass returned to the pipeline
# ---------------------------------------------------------------------------

@dataclass
class StandardsMappingResult:
    """Full result of the TSS mapping stage, ready for the pipeline."""

    # Raw LLM output
    mapping: StandardsMappingOutput

    # Mode used: "text" or "text+image"
    mode_used: str

    # Latency of the LLM call(s) in seconds
    latency_s: float

    # Whether the image fallback was triggered
    fallback_triggered: bool = False

    def to_dict(self) -> dict:
        return {
            "mode_used": self.mode_used,
            "latency_s": round(self.latency_s, 2),
            "fallback_triggered": self.fallback_triggered,
            **self.mapping.model_dump(),
        }


@dataclass
class StandardsComparisonResult:
    """Deterministic comparison between detected (from drawing) and suggested (from TSS mapping)."""

    # Norms found in both detected and suggested → already covered
    in_both: list[str] = field(default_factory=list)

    # Norms only in suggested → present these as "suggested for human validation"
    only_in_suggested: list[str] = field(default_factory=list)

    # Norms only in detected → not covered by TSS mapping → caveat
    only_in_detected: list[str] = field(default_factory=list)

    # Generic entries removed from detected before comparison
    filtered_generic: list[str] = field(default_factory=list)

    # Per-standard reasoning for the suggested ones (standard → reasoning text)
    suggested_reasoning: dict[str, str] = field(default_factory=dict)

    # Overall reasoning from the LLM mapping
    overall_reasoning: str = ""

    def to_dict(self) -> dict:
        return {
            "in_both": self.in_both,
            "only_in_suggested": self.only_in_suggested,
            "only_in_detected": self.only_in_detected,
            "filtered_generic": self.filtered_generic,
            "suggested_reasoning": self.suggested_reasoning,
            "overall_reasoning": self.overall_reasoning,
        }


# ---------------------------------------------------------------------------
# Deterministic comparison
# ---------------------------------------------------------------------------

def _normalize(norm: str) -> str:
    """Normalise a norm code for comparison: uppercase, collapse spaces."""
    return " ".join(norm.upper().split())


def compare_detected_vs_suggested(
    detected_raw: list[str],
    mapping_result: StandardsMappingResult,
) -> StandardsComparisonResult:
    """Compare detected norms (from drawing text) with TSS-mapped suggestions.

    Logic:
        1. Filter generic entries from detected (e.g. "ISO STANDARDS").
        2. Build normalised sets for both lists.
        3. Compute intersection, only-in-suggested, only-in-detected.
        4. Attach per-standard reasoning for suggested norms.

    Args:
        detected_raw:   Lista de normas extraídas do desenho (lista_normas).
        mapping_result: Output do run_tss_mapping().

    Returns:
        StandardsComparisonResult
    """
    # Step 1 — filter generics from detected
    detected_clean, filtered_generic = filter_generic_standards(detected_raw)

    # Step 2 — build normalised sets
    detected_norm_map  = {_normalize(n): n for n in detected_clean}
    suggested_norm_map = {
        _normalize(s.standard): s
        for s in mapping_result.mapping.applied_standards
    }

    detected_keys  = set(detected_norm_map.keys())
    suggested_keys = set(suggested_norm_map.keys())

    # Step 3 — classify
    in_both_keys          = detected_keys & suggested_keys
    only_in_suggested_keys = suggested_keys - detected_keys
    only_in_detected_keys  = detected_keys - suggested_keys

    in_both            = [detected_norm_map[k]          for k in sorted(in_both_keys)]
    only_in_suggested  = [suggested_norm_map[k].standard for k in sorted(only_in_suggested_keys)]
    only_in_detected   = [detected_norm_map[k]          for k in sorted(only_in_detected_keys)]

    # Step 4 — reasoning per suggested standard
    suggested_reasoning: dict[str, str] = {
        suggested_norm_map[k].standard: suggested_norm_map[k].reasoning
        for k in only_in_suggested_keys
    }

    return StandardsComparisonResult(
        in_both=in_both,
        only_in_suggested=only_in_suggested,
        only_in_detected=only_in_detected,
        filtered_generic=filtered_generic,
        suggested_reasoning=suggested_reasoning,
        overall_reasoning=mapping_result.mapping.component_inference,
    )


# ---------------------------------------------------------------------------
# Prompt builders
# ---------------------------------------------------------------------------

_SYSTEM_PROMPT = """\
You are a senior technical standards expert for Embraco/Nidec hermetic compressors.

Your task is to identify which TSS standards from the provided table apply to the CAD drawing.

## Association Rule (MANDATORY — two steps)

**Step 1 — Composition Inference:**
Analyse the title (TITLE, DOCUMENT TYPE) and drawing content to determine:
- What is the physical nature of the component? (machined part, stamped, cast, assembly, etc.)
- What material is it made of? (steel, aluminium, cast iron, copper, sintered, etc.)
- What is its function within the compressor? (individual part, sub-assembly, fully assembled product, etc.)

**Step 2 — Category Mapping:**
Based on the inference above, associate the component with the correct Applicability category.
Do NOT do literal keyword matching — apply deductive engineering reasoning.

## Example (Golden Rule)
- Title: "Stator - Stack"
- Inference: metallic steel component that forms the electric motor, individual part, not the assembled compressor
- Standard applied:  TSS 002369 (All metallic components) ✓
- Standard ignored:  TSS 001266 (Compressor assembly) — it is a part, not an assembly ✗

## Standards with Applicability "All"
TSS 002470 and TSS 002513 apply to ALL technical drawings without exception.
TSS 002420 applies to ALL components (raw materials, finished products, packaging).

## Output format
Return structured JSON according to the defined schema. All text fields must be in English.
"""


def _build_standards_context(standards: list[dict]) -> str:
    lines = ["## Available TSS Standards Table\n"]
    lines.append(f"{'Standard':<14} {'Category':<22} {'Compressor Series':<30} Applicability")
    lines.append("-" * 100)
    for s in standards:
        series = s.get("compressor_series", "All")
        lines.append(
            f"{s['standard']:<14} {s['category']:<22} {series:<30} {s['applicability']}"
        )
    return "\n".join(lines)


def _build_text_prompt(pdf_text: str, standards: list[dict]) -> str:
    ctx = _build_standards_context(standards)
    return (
        f"{_SYSTEM_PROMPT}\n\n"
        f"{ctx}\n\n"
        "## Vectorised text from the CAD drawing\n\n"
        f"{pdf_text}\n\n"
        "Based on the text above, identify the component and map all applicable standards.\n"
        "Set needs_visual_confirmation=true if the text is insufficient to be certain."
    )


def _build_image_prompt(pdf_text: str, standards: list[dict]) -> str:
    ctx = _build_standards_context(standards)
    return (
        f"{_SYSTEM_PROMPT}\n\n"
        f"{ctx}\n\n"
        "## Vectorised text from the CAD drawing (supplementary to the image)\n\n"
        f"{pdf_text}\n\n"
        "The image above shows the complete technical drawing.\n"
        "Use both the text and the image to identify the component and map all applicable standards."
    )


# ---------------------------------------------------------------------------
# LLM client — singleton
# ---------------------------------------------------------------------------

_client = None


def _get_client():
    global _client
    if _client is None:
        from google import genai
        logger.info("Initialising Gemini client (project=%s, region=%s)...", GCP_PROJECT, GCP_REGION)
        _client = genai.Client(vertexai=True, project=GCP_PROJECT, location=GCP_REGION)
        logger.info("Gemini client ready.")
    return _client


# ---------------------------------------------------------------------------
# LLM calls
# ---------------------------------------------------------------------------

def _call_text_only(
    pdf_text: str,
    standards: list[dict],
    model: str,
) -> tuple[StandardsMappingOutput, float]:
    from google.genai import types

    logger.info("[TSS MAP / text] Sending to %s...", model)
    t0 = time.time()

    prompt = _build_text_prompt(pdf_text, standards)
    response = _get_client().models.generate_content(
        model=model,
        contents=[types.Part.from_text(text=prompt)],
        config=types.GenerateContentConfig(
            response_mime_type="application/json",
            response_schema=StandardsMappingOutput,
        ),
    )
    result = StandardsMappingOutput.model_validate_json(response.text)
    latency = time.time() - t0
    logger.info(
        "[TSS MAP / text] Done in %.1fs | confidence=%.2f | fallback_needed=%s",
        latency, result.overall_confidence, result.needs_visual_confirmation,
    )
    return result, latency


def _call_text_and_image(
    pdf_text: str,
    pdf_bytes: bytes,
    standards: list[dict],
    model: str,
    dpi: int,
) -> tuple[StandardsMappingOutput, float]:
    from google.genai import types
    from src.utils.helper_func import pdf_to_images_base64

    logger.info("[TSS MAP / text+image] Rendering PDF and sending to %s...", model)
    t0 = time.time()

    images_b64 = pdf_to_images_base64(pdf_bytes, dpi=dpi)
    prompt = _build_image_prompt(pdf_text, standards)

    parts: list = []
    for i, img_b64 in enumerate(images_b64):
        parts.append(types.Part.from_bytes(data=base64.b64decode(img_b64), mime_type="image/png"))
        parts.append(types.Part.from_text(text=f"[Page {i + 1} of the drawing above]"))
    parts.append(types.Part.from_text(text=prompt))

    response = _get_client().models.generate_content(
        model=model,
        contents=parts,
        config=types.GenerateContentConfig(
            response_mime_type="application/json",
            response_schema=StandardsMappingOutput,
        ),
    )
    result = StandardsMappingOutput.model_validate_json(response.text)
    latency = time.time() - t0
    logger.info(
        "[TSS MAP / text+image] Done in %.1fs | confidence=%.2f",
        latency, result.overall_confidence,
    )
    return result, latency


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def run_tss_mapping(
    pdf_bytes: bytes,
    pdf_text: str,
    *,
    normas_path: Path = DEFAULT_NORMAS_PATH,
    model: str = MODEL_ID,
    confidence_threshold: float = CONFIDENCE_THRESHOLD,
    dpi: int = 150,
) -> StandardsMappingResult:
    """Run TSS standards mapping for a CAD drawing.

    Tries text-only first. If confidence < threshold or the LLM itself signals
    needs_visual_confirmation, falls back to text + rendered image.

    Args:
        pdf_bytes:            Raw bytes of the revised PDF.
        pdf_text:             Pre-extracted vectorised text (all pages joined).
        normas_path:          Path to normas.xlsx.
        model:                Gemini model ID.
        confidence_threshold: Minimum confidence to accept text-only result.
        dpi:                  DPI for image rendering on fallback.

    Returns:
        StandardsMappingResult
    """
    standards = load_standards_table(normas_path)

    # --- text-only call ---
    mapping, latency = _call_text_only(pdf_text, standards, model)

    fallback = (
        mapping.overall_confidence < confidence_threshold
        or mapping.needs_visual_confirmation
    )

    if fallback:
        reasons = []
        if mapping.overall_confidence < confidence_threshold:
            reasons.append(
                f"confidence={mapping.overall_confidence:.2f} < threshold={confidence_threshold:.2f}"
            )
        if mapping.needs_visual_confirmation:
            reasons.append("LLM requested visual confirmation")
        logger.warning("[TSS MAP] Fallback triggered: %s", " | ".join(reasons))

        mapping, img_latency = _call_text_and_image(pdf_text, pdf_bytes, standards, model, dpi)
        latency += img_latency

    return StandardsMappingResult(
        mapping=mapping,
        mode_used="text+image" if fallback else "text",
        latency_s=latency,
        fallback_triggered=fallback,
    )
