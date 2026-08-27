"""Generate validation spreadsheets from integrated review JSON results.

Reads each scripts/results/test_<N>/integrated_review.json, counts GD&T,
Datums and Cotas, then writes the values into a copy of the validation
template at scripts/validation/41-50 Structured reviews/.

Output copies go to scripts/results/test_<N>/validation_filled.xlsx

Usage:
    python scripts/generate_validation.py
    python scripts/generate_validation.py --tests 41 42 43
"""

from __future__ import annotations

import argparse
from copy import copy
import json
import logging
import shutil
import sys
from pathlib import Path

# Ensure project root is on sys.path so src.metrics can be imported
# regardless of where the script is called from
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

RESULTS_DIR    = Path("scripts/results")
VALIDATION_DIR = Path("scripts/validation/41-50 Structured reviews")

OBJECTIVE_METRICS = [
    ("Quantidade de cotas", "Quantidade de cotas presentes no desenho inteiro"),
    ("Quantidade de cotas HIC", "Quantidade de cotas classificadas como HIC (▽)"),
    ("Quantidade de cotas CTQ", "Quantidade de cotas classificadas como CTQ (▼)"),
    ("Quantidade de cotas CTQ-S", "Quantidade de cotas classificadas como CTQ-S (⊕)"),
    ("Quantidade de GD&Ts", "Quantidade de GD&Ts presentes no desenho inteiro"),
    ("Quantidade de Datums Reference", "Quantidade de datum feature indicators encontrados"),
    ("Lista de datums reference", "Lista única dos datums encontrados"),
    ("Quantidade de revisões", "Quantidade de revisões preenchidas na tabela de revisões"),
    ("Quantidade de notas", "Quantidade de itens numerados na lista de notas"),
    ("Quantidade de códigos", "Quantidade de códigos na tabela de materiais/componentes"),
]

# ---------------------------------------------------------------------------
# Helpers: find template for each test number
# ---------------------------------------------------------------------------

def _template_path(n: int) -> Path:
    """Return path to the validation template for test N."""
    candidates = [
        VALIDATION_DIR / f"Drawing Data Extraction - Number_ {n}.xlsx",
        VALIDATION_DIR / f"Drawing Data Extraction - Number_ {n}_.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError(
        f"No validation template found for test {n} in {VALIDATION_DIR}"
    )


# ---------------------------------------------------------------------------
# Label -> row mapping builder
# ---------------------------------------------------------------------------

def _build_label_row_map(ws) -> dict[str, int]:
    """Map cell label (col A text) to its row number in the sheet."""
    mapping: dict[str, int] = {}
    for row in ws.iter_rows():
        cell_a = row[0]
        if cell_a.value and isinstance(cell_a.value, str):
            mapping[cell_a.value.strip()] = cell_a.row
    return mapping


def _set_value(ws, label_map: dict[str, int], label: str, value) -> bool:
    """Write value into column B of the row matching label. Returns True if found."""
    row = label_map.get(label)
    if row is None:
        return False
    ws.cell(row=row, column=2, value=value)
    return True


def _spreadsheet_value(value):
    """Convert structured JSON values to an Excel-compatible scalar."""

    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(item) for item in value) or None
    if isinstance(value, dict):
        return "; ".join(f"{key}: {item}" for key, item in value.items()) or None
    return value


def _ensure_objective_metric_rows(ws) -> None:
    """Ensure the generated workbook contains every customer metric row."""

    label_map = _build_label_row_map(ws)
    objective_row = label_map.get("Objective Metrics")
    references_row = label_map.get("References")
    if objective_row is None or references_row is None:
        logger.warning("Objective Metrics or References section not found in template.")
        return

    existing_values = {
        label: ws.cell(row=row, column=2).value
        for label, row in label_map.items()
        if label in {item[0] for item in OBJECTIVE_METRICS}
    }
    current_slots = max(0, references_row - objective_row - 1)
    rows_to_add = max(0, len(OBJECTIVE_METRICS) - current_slots)

    if rows_to_add:
        overlapping_merges = [
            str(merged)
            for merged in ws.merged_cells.ranges
            if merged.min_row <= references_row <= merged.max_row
        ]
        for merged in overlapping_merges:
            ws.unmerge_cells(merged)
        ws.insert_rows(references_row, amount=rows_to_add)
        references_row += rows_to_add
        if overlapping_merges:
            ws.merge_cells(
                start_row=references_row,
                start_column=1,
                end_row=references_row,
                end_column=2,
            )

    style_source_row = objective_row + 1
    for offset, (label, description) in enumerate(OBJECTIVE_METRICS, start=1):
        target_row = objective_row + offset
        for column in range(1, 4):
            source = ws.cell(row=style_source_row, column=column)
            target = ws.cell(row=target_row, column=column)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
        ws.cell(row=target_row, column=1, value=label)
        ws.cell(row=target_row, column=2, value=existing_values.get(label))
        ws.cell(row=target_row, column=3, value=description)


# ---------------------------------------------------------------------------
# Data extraction from integrated_review.json
# ---------------------------------------------------------------------------

def _extract_data(review: dict, cotas_result: dict, pdf_path: Path) -> dict:
    """Extract all fillable fields from review JSON + counters."""
    pc = review.get("part_classification", {})
    header        = pc.get("header", {}) or {}
    drawing_block = pc.get("drawing_block", {}) or {}

    # GD&T and datums from review JSON
    from src.metrics.counters import count_gdt_and_datums
    gdt_counts = count_gdt_and_datums(review)
    objective_metrics = review.get("objective_metrics", {}) or {}

    return {
        "Compressor Series Code": header.get("compressor_series_code"),
        "Applicable norms": "; ".join(
            str(n) for n in pc.get("lista_normas", []) or []
        ) or None,
        "Last revision date": header.get("last_revision_date"),
        "Materials":               drawing_block.get("materials"),
        "Code":                    drawing_block.get("material_code"),
        "Material Code":           drawing_block.get("material_code"),
        "Drawn by":                drawing_block.get("drawn_by"),
        "Approved by":             drawing_block.get("approved_by"),
        "Drawing Code (ECM)":      drawing_block.get("drawing_code_ecm"),
        "Date":                    drawing_block.get("date"),
        "Name and document type":  drawing_block.get("name_and_document_type"),
        "General tolerance":       drawing_block.get("general_tolerance"),
        "Angular tolerance":       drawing_block.get("angular_tolerance"),
        "Scale":                   drawing_block.get("scale"),
        "Unit":                    drawing_block.get("unit"),
        "Replace":                 drawing_block.get("replace"),
        "Number":                  drawing_block.get("number"),
        "Quantidade de cotas": objective_metrics.get(
            "Quantidade de cotas", cotas_result.get("total_cotas", 0)
        ),
        "Quantidade de cotas HIC": objective_metrics.get(
            "Quantidade de cotas HIC", pc.get("quantidade_cotas_hic")
        ),
        "Quantidade de cotas CTQ": objective_metrics.get(
            "Quantidade de cotas CTQ", pc.get("quantidade_cotas_ctq")
        ),
        "Quantidade de cotas CTQ-S": objective_metrics.get(
            "Quantidade de cotas CTQ-S", pc.get("quantidade_cotas_ctq_s")
        ),
        "Quantidade de GD&Ts": objective_metrics.get(
            "Quantidade de GD&Ts", gdt_counts["total_gdts"]
        ),
        "Quantidade de Datums Reference": objective_metrics.get(
            "Quantidade de Datums Reference", gdt_counts["total_datums"]
        ),
        "Lista de datums reference": _spreadsheet_value(
            objective_metrics.get(
                "Lista de datums reference", _extract_datum_list(review)
            )
        ),
        "Quantidade de revisões": objective_metrics.get(
            "Quantidade de revisões", pc.get("quantidade_revisoes")
        ),
        "Quantidade de notas": objective_metrics.get(
            "Quantidade de notas", pc.get("quantidade_notas")
        ),
        "Quantidade de códigos": objective_metrics.get(
            "Quantidade de códigos", pc.get("quantidade_codigos")
        ),
    }


def _extract_datum_list(review: dict) -> str | None:
    """Extract unique datum labels from GD&T pages."""
    datums: list[str] = []
    for page in review.get("gdt", []):
        for defn in page.get("datum_definitions", []) or []:
            label = defn.get("label") or defn.get("datum_label")
            if label and label not in datums:
                datums.append(str(label))
    return ", ".join(datums) if datums else None


# ---------------------------------------------------------------------------
# Sheet filler
# ---------------------------------------------------------------------------

def fill_single_drawing_sheet(ws, data: dict) -> None:
    """Fill the 'Single Drawing Data Extraction' sheet with extracted data."""
    label_map = _build_label_row_map(ws)

    for label, value in data.items():
        if value is None:
            continue
        found = _set_value(ws, label_map, label, value)
        if not found:
            logger.debug("Label not found in sheet: %r", label)


def fill_comparison_sheet(ws, review: dict) -> None:
    """Fill the 'Two Drawings Comparison' sheet with detected changes."""
    # Find the header row (row with 'Change ID')
    header_row = None
    for row in ws.iter_rows(max_row=20):
        if any(cell.value == "Change ID" for cell in row):
            header_row = row[0].row
            break

    if header_row is None:
        logger.warning("Could not find 'Change ID' header in Two Drawings Comparison sheet.")
        return

    # Clear existing data rows below header
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    data_row       = header_row + 1
    change_counter = 1

    # True changes from LLM comparison pages
    for page in review.get("comparison", {}).get("pages", []) or []:
        # true_changes is a list of dicts in the serialised JSON
        for change in page.get("true_changes", []) or []:
            if isinstance(change, dict):
                description = change.get("description", "")
            else:
                description = str(change)

            ws.cell(row=data_row, column=1, value=change_counter)
            ws.cell(row=data_row, column=2, value=str(description))
            # Col 3: Previous Quadrant — left blank for human fill
            # Col 4: Current Quadrant  — left blank for human fill
            # Col 5: Sheet Format      — filled below from paper_format_changes
            data_row       += 1
            change_counter += 1

    # Deterministic paper format changes
    for fmt_change in review.get("comparison", {}).get("paper_format_changes", []) or []:
        desc = fmt_change.get("description", "Format change")
        sheet_fmt = fmt_change.get("original", {}).get("size") or ""
        ws.cell(row=data_row, column=1, value=change_counter)
        ws.cell(row=data_row, column=2, value=str(desc))
        ws.cell(row=data_row, column=5, value=str(sheet_fmt) if sheet_fmt else None)
        data_row       += 1
        change_counter += 1


# ---------------------------------------------------------------------------
# Main per-test function
# ---------------------------------------------------------------------------

def process_test(n: int) -> bool:
    """Process a single test: read JSON, count cotas, write filled xlsx."""
    import openpyxl
    from src.metrics.counters import count_cotas

    result_dir = RESULTS_DIR / f"test_{n}"
    json_path  = result_dir / "integrated_review.json"

    if not json_path.exists():
        logger.warning("[test %d] JSON not found: %s — skipping.", n, json_path)
        return False

    # Find revised PDF (draw_2) for cotas counting
    base = Path("CAD_Review_Test_Battery_V1/2. Comparison Analysis") / str(n)
    revised_pdfs = sorted(base.glob("*draw_2.pdf"))
    if not revised_pdfs:
        revised_pdfs = sorted(base.glob("*_draw_*.pdf"))
    revised_pdf = revised_pdfs[-1] if revised_pdfs else None

    # Load review JSON
    with open(json_path, encoding="utf-8") as f:
        review = json.load(f)

    # Count cotas
    if revised_pdf and revised_pdf.exists():
        logger.info("[test %d] Counting cotas in: %s", n, revised_pdf.name)
        cotas_result = count_cotas(revised_pdf)
        logger.info("[test %d] Cotas found: %d", n, cotas_result["total_cotas"])
    else:
        logger.warning("[test %d] Revised PDF not found for cotas count.", n)
        cotas_result = {"total_cotas": 0, "per_page": [], "cotas_detail": []}

    # Load template
    try:
        template = _template_path(n)
    except FileNotFoundError as e:
        logger.error("[test %d] %s", n, e)
        return False

    output_path = result_dir / "validation_filled.xlsx"
    shutil.copy2(template, output_path)

    wb = openpyxl.load_workbook(output_path)

    # Fill Single Drawing Data Extraction
    if "Single Drawing Data Extraction" in wb.sheetnames:
        _ensure_objective_metric_rows(wb["Single Drawing Data Extraction"])
        data = _extract_data(review, cotas_result, revised_pdf or Path())
        fill_single_drawing_sheet(wb["Single Drawing Data Extraction"], data)
    else:
        logger.warning("[test %d] Sheet 'Single Drawing Data Extraction' not found.", n)

    # Fill Two Drawings Comparison
    if "Two Drawings Comparison" in wb.sheetnames:
        fill_comparison_sheet(wb["Two Drawings Comparison"], review)
    else:
        logger.warning("[test %d] Sheet 'Two Drawings Comparison' not found.", n)

    wb.save(output_path)
    logger.info("[test %d] Validation written -> %s", n, output_path)
    return True


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%H:%M:%S",
    )

    parser = argparse.ArgumentParser(
        description="Generate filled validation sheets from review JSON results"
    )
    parser.add_argument(
        "--tests",
        type=int,
        nargs="+",
        default=list(range(41, 51)),
        help="Test numbers to process (default: all 41-50)",
    )
    args = parser.parse_args()

    passed = 0
    failed = []

    for n in args.tests:
        ok = process_test(n)
        if ok:
            passed += 1
        else:
            failed.append(n)

    print(f"\nDone: {passed}/{len(args.tests)} tests processed.")
    if failed:
        print(f"Skipped (no JSON yet): {failed}")


if __name__ == "__main__":
    main()
