import openpyxl
import os

base_dir = r'c:\Users\madeinweb\Documents\GitHub\projeto-nidec-cad-review\nidec-cad-review\scripts\validation\41-50 Structured reviews'

files = [
    'Drawing Data Extraction - Number_ 41.xlsx',
    'Drawing Data Extraction - Number_ 42.xlsx',
    'Drawing Data Extraction - Number_ 43_.xlsx',
    'Drawing Data Extraction - Number_ 44.xlsx',
    'Drawing Data Extraction - Number_ 45.xlsx',
    'Drawing Data Extraction - Number_ 46.xlsx',
    'Drawing Data Extraction - Number_ 47.xlsx',
    'Drawing Data Extraction - Number_ 48.xlsx',
    'Drawing Data Extraction - Number_ 49.xlsx',
    'Drawing Data Extraction - Number_ 50.xlsx',
]

for fname in files:
    fpath = os.path.join(base_dir, fname)
    wb = openpyxl.load_workbook(fpath)
    ws = wb['Single Drawing Data Extraction']
    print(f'\n=== {fname} ===')
    print(f'Dimensions: {ws.dimensions}')
    # Find Objective Metrics row and print from there
    obj_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 'Objective Metrics':
                obj_row = cell.row
                break
        if obj_row:
            break
    
    if obj_row:
        print(f'Objective Metrics starts at row {obj_row}')
        print('Rows from Objective Metrics onward:')
        for r in ws.iter_rows(min_row=obj_row):
            row_vals = [(cell.column_letter, cell.value) for cell in r if cell.value is not None]
            if row_vals:
                print(f'  Row {r[0].row}: {row_vals}')
    else:
        print('WARNING: Objective Metrics NOT FOUND')
