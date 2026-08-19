import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Alignment

fpath = r'c:\Users\madeinweb\Documents\GitHub\projeto-nidec-cad-review\nidec-cad-review\scripts\validation\41-50 Structured reviews\Drawing Data Extraction - Number_ 43_.xlsx'

wb = openpyxl.load_workbook(fpath)
ws = wb['Single Drawing Data Extraction']

print('Styles for rows 20-33:')
for row in ws.iter_rows(min_row=20, max_row=33):
    for cell in row:
        print(f'  {cell.coordinate}: value={repr(cell.value)}')
        print(f'    font bold={cell.font.bold}, color={cell.font.color.rgb if cell.font.color.type == "rgb" else cell.font.color.type}')
        if cell.fill.fill_type != 'none' and cell.fill.fill_type is not None:
            try:
                print(f'    fill type={cell.fill.fill_type}, fgColor={cell.fill.fgColor.rgb}')
            except:
                print(f'    fill type={cell.fill.fill_type}')
        print(f'    alignment: horizontal={cell.alignment.horizontal}, vertical={cell.alignment.vertical}')
