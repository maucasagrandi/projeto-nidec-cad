"""
Standardize the Objective Metrics section in validation spreadsheets 41-50.

Target structure (starting at the "Objective Metrics" header row):
  Row obj+0: Objective Metrics | Human Answer | Description   <- header (bold, gray fill)
  Row obj+1: Quantidade de cotas | (empty) | Quantidade de cotas presentes no desenho inteiro
  Row obj+2: Quantidade de GD&Ts | (empty) | Quantidade de GD&Ts presentes no desenho inteiro
  Row obj+3: Quantidade de revisões | (empty) | Quantidade de revisões já realizadas no desenho (pela tabela de revisões)
  Row obj+4: Quantidade de notas | (empty) | Quantidade de notas na lista de notas (NOTES: 1 - ) no desenho inteiro
  Row obj+5: Quantidade de códigos | (empty) | Quantidades de códigos na tabela (por exemplo A, B, C, D, I, L, #)
  Row obj+6: References (A:B merged) | | Description    <- header (bold, gray fill)
  Row obj+7: Norms table | Do not fill | Norms table used as reference to model (CAD Review - Standards)
  Row obj+8: CAD Review model version | Do not fill | CAD Review model version, for the delivery of August 31st 2026, it will be v1
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border

BASE_DIR = r'c:\Users\madeinweb\Documents\GitHub\projeto-nidec-cad-review\nidec-cad-review\scripts\validation\41-50 Structured reviews'

FILES = [
    'Drawing Data Extraction - Number_ 41.xlsx',
    'Drawing Data Extraction - Number_ 42.xlsx',
    'Drawing Data Extraction - Number_ 43_.xlsx',
    'Drawing Data Extraction - Number_ 44.xlsx',
    'Drawing Data Extraction - Number_ 45.xlsx',
    'Drawing Data Extraction - Number_ 46.xlsx',
    'Drawing Data Extraction - Number_ 47.xlsx',
    'Drawing Data Extraction - Number_ 48.xlsx',
    'Drawing Data Extraction - Number_ 49.xlsx',
    'Drawing Data Extraction - Number_ 50.xlsx',
]

OBJECTIVE_METRICS_ROWS = [
    ('Quantidade de cotas',      None, 'Quantidade de cotas presentes no desenho inteiro'),
    ('Quantidade de GD&Ts',      None, 'Quantidade de GD&Ts presentes no desenho inteiro'),
    ('Quantidade de revisões',   None, 'Quantidade de revisões já realizadas no desenho (pela tabela de revisões)'),
    ('Quantidade de notas',      None, 'Quantidade de notas na lista de notas (NOTES: 1 - ) no desenho inteiro'),
    ('Quantidade de códigos',    None, 'Quantidades de códigos na tabela (por exemplo A, B, C, D, I, L, #)'),
]

REFERENCES_ROWS = [
    ('Norms table',              'Do not fill', 'Norms table used as reference to model (CAD Review - Standards)'),
    ('CAD Review model version', 'Do not fill', 'CAD Review model version, for the delivery of August 31st 2026, it will be v1'),
]

GRAY_FILL = PatternFill(fill_type='solid', fgColor='D9D9D9')
NO_FILL   = PatternFill(fill_type=None)
RED_FONT  = Font(color='FF0000')
BOLD_FONT = Font(bold=True)
NORMAL_FONT = Font()
CENTER_ALIGN = Alignment(horizontal='center')
LEFT_BOTTOM_ALIGN = Alignment(horizontal=None, vertical='bottom')
CENTER_BOTTOM_ALIGN = Alignment(horizontal='center', vertical='bottom')
NO_BORDER = Border()


def set_cell(cell, value, font=None, fill=None, alignment=None, border=None):
    """Safely set a normal (non-merged) cell's value and style."""
    cell.value = value
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border


def clear_cell(cell):
    """Clear value and reset styles on a normal (non-merged) cell."""
    cell.value = None
    cell.font = NORMAL_FONT
    cell.fill = NO_FILL
    cell.alignment = Alignment()
    cell.border = NO_BORDER


def unmerge_all_in_range(ws, min_row, max_row):
    """
    Remove any merged cell ranges that overlap rows min_row..max_row.
    Returns list of removed ranges (as strings) so we can re-apply if needed.
    """
    to_remove = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= min_row and mr.max_row <= max_row:
            to_remove.append(str(mr))
    for r in to_remove:
        ws.unmerge_cells(r)
    return to_remove


def standardize_file(fpath):
    fname = os.path.basename(fpath)
    print(f'\nProcessing: {fname}')
    wb = openpyxl.load_workbook(fpath)
    ws = wb['Single Drawing Data Extraction']

    # 1. Find the "Objective Metrics" header row
    obj_row_idx = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 'Objective Metrics':
                obj_row_idx = cell.row
                break
        if obj_row_idx is not None:
            break

    if obj_row_idx is None:
        print(f'  WARNING: "Objective Metrics" not found! Skipping.')
        return

    print(f'  Objective Metrics header at row {obj_row_idx}')

    # 2. Determine the current last row with data
    max_data_row = obj_row_idx
    for row in ws.iter_rows(min_row=obj_row_idx):
        for cell in row:
            if cell.value is not None:
                max_data_row = max(max_data_row, cell.row)

    # Target rows:
    # obj_row_idx + 1..5 = 5 metric data rows
    # obj_row_idx + 6    = References header
    # obj_row_idx + 7..8 = 2 reference data rows
    target_last_row = obj_row_idx + 8

    # 3. Unmerge any merged cells in rows obj_row_idx+1 onwards
    #    (the References header row merge will be re-applied at the correct row)
    removed_merges = unmerge_all_in_range(ws, obj_row_idx + 1, max(max_data_row, target_last_row))
    if removed_merges:
        print(f'  Removed merged ranges: {removed_merges}')

    # 4. Write the 5 Objective Metrics data rows
    for i, (a_val, b_val, c_val) in enumerate(OBJECTIVE_METRICS_ROWS):
        r = obj_row_idx + 1 + i
        set_cell(ws.cell(row=r, column=1), a_val,  font=NORMAL_FONT, fill=NO_FILL, alignment=LEFT_BOTTOM_ALIGN)
        set_cell(ws.cell(row=r, column=2), b_val,  font=NORMAL_FONT, fill=NO_FILL, alignment=CENTER_BOTTOM_ALIGN)
        set_cell(ws.cell(row=r, column=3), c_val,  font=NORMAL_FONT, fill=NO_FILL, alignment=LEFT_BOTTOM_ALIGN)

    # 5. Write References header row (A:B merged, bold, gray)
    ref_header_row = obj_row_idx + 6
    # Write value in A (the master cell of the merge), then merge A:B
    set_cell(ws.cell(row=ref_header_row, column=1), 'References', font=BOLD_FONT, fill=GRAY_FILL, alignment=CENTER_ALIGN)
    set_cell(ws.cell(row=ref_header_row, column=2), None, font=NORMAL_FONT, fill=NO_FILL, alignment=Alignment())
    set_cell(ws.cell(row=ref_header_row, column=3), 'Description', font=BOLD_FONT, fill=GRAY_FILL, alignment=CENTER_ALIGN)
    # Re-apply merge A:B on this row
    ws.merge_cells(start_row=ref_header_row, start_column=1, end_row=ref_header_row, end_column=2)
    print(f'  Re-merged A{ref_header_row}:B{ref_header_row}')

    # 6. Write the 2 References data rows
    for i, (a_val, b_val, c_val) in enumerate(REFERENCES_ROWS):
        r = obj_row_idx + 7 + i
        set_cell(ws.cell(row=r, column=1), a_val,  font=NORMAL_FONT,  fill=NO_FILL, alignment=CENTER_ALIGN)
        set_cell(ws.cell(row=r, column=2), b_val,  font=RED_FONT,     fill=NO_FILL, alignment=CENTER_ALIGN)
        set_cell(ws.cell(row=r, column=3), c_val,  font=NORMAL_FONT,  fill=NO_FILL, alignment=Alignment())

    # 7. Clear any leftover rows beyond target_last_row
    if max_data_row > target_last_row:
        print(f'  Clearing leftover rows {target_last_row + 1}..{max_data_row}')
        for r in range(target_last_row + 1, max_data_row + 1):
            for col in range(1, 4):
                clear_cell(ws.cell(row=r, column=col))

    wb.save(fpath)
    print(f'  Saved successfully.')


def verify_file(fpath):
    wb = openpyxl.load_workbook(fpath)
    ws = wb['Single Drawing Data Extraction']

    obj_row_idx = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 'Objective Metrics':
                obj_row_idx = cell.row
                break
        if obj_row_idx is not None:
            break

    print(f'\n  {os.path.basename(fpath)}')
    if obj_row_idx is None:
        print('    ERROR: Objective Metrics not found!')
        return

    for row in ws.iter_rows(min_row=obj_row_idx, max_row=obj_row_idx + 9):
        vals = {cell.column_letter: cell.value for cell in row if cell.value is not None}
        all_none = all(cell.value is None for cell in row)
        if not all_none:
            print(f'    Row {row[0].row}: A={repr(vals.get("A"))}  B={repr(vals.get("B"))}  C={repr(vals.get("C"))}')

    # Check for any extra data rows beyond target
    extra_rows = []
    for row in ws.iter_rows(min_row=obj_row_idx + 9):
        for cell in row:
            if cell.value is not None:
                extra_rows.append(f'row {cell.row} col {cell.column_letter}: {repr(cell.value)}')
    if extra_rows:
        print(f'    EXTRA DATA FOUND: {extra_rows}')
    else:
        print(f'    No extra data beyond row {obj_row_idx + 8}. OK.')

    # Check merges
    print(f'    Merged ranges: {list(ws.merged_cells.ranges)}')


if __name__ == '__main__':
    for fname in FILES:
        fpath = os.path.join(BASE_DIR, fname)
        if not os.path.exists(fpath):
            print(f'File not found: {fpath}')
            continue
        standardize_file(fpath)

    print('\n' + '=' * 70)
    print('VERIFICATION')
    print('=' * 70)
    for fname in FILES:
        fpath = os.path.join(BASE_DIR, fname)
        if os.path.exists(fpath):
            verify_file(fpath)
