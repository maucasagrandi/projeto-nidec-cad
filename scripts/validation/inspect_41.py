import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\madeinweb\Documents\GitHub\projeto-nidec-cad-review\nidec-cad-review\scripts\validation\41-50 Structured reviews\Drawing Data Extraction - Number_ 41.xlsx')
print('Sheets:', wb.sheetnames)
ws = wb['Single Drawing Data Extraction']
print('Sheet dimensions:', ws.dimensions)
print()
print('All non-empty rows:')
for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            print(f'  Row {cell.row}, Col {cell.column} ({cell.column_letter}): {repr(cell.value)}')
