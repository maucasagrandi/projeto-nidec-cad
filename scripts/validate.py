"""Validate model outputs against human answers from validation spreadsheets.

Reads:
    scripts/results/test_<N>/validation_filled.xlsx  — model answers (col B)
    scripts/validation/41-50 Structured reviews/Drawing Data Extraction - Number_ <N>.xlsx
                                                     — human answers (col B)

Scoring:
    - Textual fields  → LLM-as-a-Judge  (0.0–1.0 semantic similarity)
    - Numeric metrics → MAPE-based score (0.0–1.0, human is ground truth)
    - Changes (Two Drawings Comparison) → LLM-as-a-Judge per change description

Output:
    scripts/results/test_<N>/validation_report.json  — per-field scores
    scripts/results/validation_summary.json          — score per test + overall

Usage:
    python scripts/validate.py
    python scripts/validate.py --tests 41 42 43
    python scripts/validate.py --no-llm   # skip LLM, only MAPE
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import time
from pathlib import Path
from typing import Any

_PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

RESULTS_DIR    = Path("scripts/results")
VALIDATION_DIR = Path("scripts/validation/41-50 Structured reviews")

# ---------------------------------------------------------------------------
# Score weights for the composite score
# ---------------------------------------------------------------------------

WEIGHT_DRAWING_BLOCK = 0.40   # header + drawing block fields
WEIGHT_METRICS       = 0.35   # objective metrics (MAPE)
WEIGHT_CHANGES       = 0.25   # two drawings comparison

# ---------------------------------------------------------------------------
# Field classification
# ---------------------------------------------------------------------------

# These fields are evaluated by LLM (semantic similarity)
TEXTUAL_FIELDS = [
    "Last revision date",
    "Materials",
    "Code",
    "Material Code",
    "Drawn by",
    "Approved by",
    "Drawing Code (ECM)",
    "Date",
    "Name and document type",
    "General tolerance",
    "Angular tolerance",
    "Scale",
    "Unit",
    "Replace",
]

# These fields use exact string match (numbers, codes)
EXACT_FIELDS = [
    "Number",
]

# These fields use MAPE score (numeric, human is ground truth)
METRIC_FIELDS = [
    "Quantidade de cotas",
    "Quantidade de GD&Ts",
    "Quantidade de Datums Reference",
    "Quantidade de revisões",
    "Quantidade de notas",
    "Quantidade de códigos",
]

# Fields to skip entirely
SKIP_FIELDS = {
    "Norms table", "CAD Review model version",
    "Drawing Number in Test Battery 1",
    "Header", "Drawing Block Data", "Objective Metrics", "References",
    "Compressor Series Code", "Applicable norms",
    "Lista de datums reference",
    "Quantidade de cotas HIC", "Quantidade de cotas CTQ", "Quantidade de cotas CTQ-S",
}

# ---------------------------------------------------------------------------
# MAPE-based score (0 to 1, human is ground truth)
# ---------------------------------------------------------------------------

def mape_score(human: float, model: float) -> float:
    """Return 1 - MAPE, clamped to [0, 1].

    Formula: score = max(0, 1 - |model - human| / human)
    If human == 0 and model == 0 → 1.0 (both agree on zero)
    If human == 0 and model != 0 → 0.0 (model invented counts)
    """
    if human == 0 and model == 0:
        return 1.0
    if human == 0:
        return 0.0
    raw = 1.0 - abs(model - human) / human
    return max(0.0, raw)


# ---------------------------------------------------------------------------
# Exact match score
# ---------------------------------------------------------------------------

def exact_score(human: str, model: str) -> float:
    """Normalised exact match: 1.0 if equal (case-insensitive strip), else 0.0."""
    if human is None and model is None:
        return 1.0
    if human is None or model is None:
        return 0.0
    return 1.0 if str(human).strip().lower() == str(model).strip().lower() else 0.0


# ---------------------------------------------------------------------------
# LLM-as-a-Judge
# ---------------------------------------------------------------------------

_llm_client = None


def _get_llm_client():
    global _llm_client
    if _llm_client is None:
        from dotenv import load_dotenv
        load_dotenv()
        from google import genai
        project = os.getenv("GCP_PROJECT_ID", "acim-global-data-lake-sandbox")
        region  = os.getenv("GCP_REGION", "us-east5")
        _llm_client = genai.Client(vertexai=True, project=project, location=region)
        logger.info("Gemini client initialised (project=%s, region=%s)", project, region)
    return _llm_client


_JUDGE_SYSTEM = """You are a strict but fair evaluator comparing two text values
from a CAD technical drawing review system.

Your task: decide how semantically equivalent the Model Answer is to the Human Answer.
The Human Answer is the ground truth provided by a domain expert.

Return ONLY a valid JSON object with exactly two fields:
{
  "score": <float between 0.0 and 1.0>,
  "reason": "<one sentence explanation>"
}

Scoring guide:
  1.0  — identical or trivially equivalent (different spacing/punctuation, same meaning)
  0.9  — same information, minor formatting difference (e.g. "+-0,2" vs "±0.2")
  0.8  — same core value, minor extra/missing detail
  0.6  — partially correct (e.g. right name but wrong format, or only part of the value)
  0.3  — related but wrong (e.g. different date, different name)
  0.0  — completely wrong or absent when an answer was expected

Special rules:
- "Empty" in Human Answer means the field was blank in the drawing.
  If Model Answer is null, empty, or "Empty", score = 1.0.
  If Model Answer has a value when Human says "Empty", score = 0.0.
- null / None / empty string on both sides → score = 1.0
- If only one side is null/empty and the other has a value → score = 0.0
"""


def llm_judge(
    field: str,
    human: str | None,
    model: str | None,
    model_id: str = "gemini-2.5-flash",
    retries: int = 2,
) -> dict[str, Any]:
    """Ask LLM to score semantic equivalence between human and model values.

    Returns dict with keys: score (float), reason (str), method (str).
    Falls back to exact_score on LLM failure.
    """
    from google.genai import types
    from pydantic import BaseModel, Field

    class JudgeOutput(BaseModel):
        score:  float = Field(ge=0.0, le=1.0)
        reason: str

    human_str = str(human) if human is not None else "null"
    model_str = str(model) if model is not None else "null"

    prompt = (
        f"{_JUDGE_SYSTEM}\n\n"
        f"Field: {field}\n"
        f"Human Answer: {human_str}\n"
        f"Model Answer: {model_str}\n"
    )

    for attempt in range(retries + 1):
        try:
            client = _get_llm_client()
            response = client.models.generate_content(
                model=model_id,
                contents=[types.Part.from_text(text=prompt)],
                config=types.GenerateContentConfig(
                    response_mime_type="application/json",
                    response_schema=JudgeOutput,
                ),
            )
            result = JudgeOutput.model_validate_json(response.text)
            return {
                "score":  round(result.score, 4),
                "reason": result.reason,
                "method": "llm",
            }
        except Exception as exc:
            logger.warning(
                "LLM judge attempt %d/%d failed for field %r: %s",
                attempt + 1, retries + 1, field, exc,
            )
            if attempt < retries:
                time.sleep(2)

    # Fallback to exact match
    sc = exact_score(human_str, model_str)
    return {
        "score":  sc,
        "reason": "LLM unavailable — fell back to exact match",
        "method": "exact_fallback",
    }


# ---------------------------------------------------------------------------
# LLM judge for change descriptions (Two Drawings Comparison)
# ---------------------------------------------------------------------------

_CHANGES_JUDGE_SYSTEM = """You are evaluating a CAD drawing comparison system.

Given a list of human-annotated changes (ground truth) and a list of model-detected
changes, score how well the model captured the changes.

Return ONLY valid JSON:
{
  "score": <float 0.0–1.0>,
  "matched": <int — number of human changes the model correctly detected>,
  "total_human": <int — total human changes>,
  "reason": "<one sentence summary>"
}

Scoring:
  1.0 — model found all changes with accurate descriptions
  0.8 — model found most changes, minor description differences
  0.6 — model found about half the changes
  0.3 — model found few changes or descriptions are mostly wrong
  0.0 — model found no relevant changes

A change is "correctly detected" if the description semantically matches,
even if wording differs. Ignore quadrant location differences.
"""


def llm_judge_changes(
    human_changes: list[str],
    model_changes: list[str],
    model_id: str = "gemini-2.5-flash",
    retries: int = 2,
) -> dict[str, Any]:
    """Score model-detected changes vs human ground truth changes."""
    from google.genai import types
    from pydantic import BaseModel, Field

    class ChangesOutput(BaseModel):
        score:        float = Field(ge=0.0, le=1.0)
        matched:      int
        total_human:  int
        reason:       str

    if not human_changes and not model_changes:
        return {"score": 1.0, "matched": 0, "total_human": 0,
                "reason": "No changes expected or detected.", "method": "trivial"}

    if not human_changes:
        return {"score": 0.0, "matched": 0, "total_human": 0,
                "reason": "No human changes but model detected some.", "method": "trivial"}

    human_list = "\n".join(f"  {i+1}. {c}" for i, c in enumerate(human_changes))
    model_list = "\n".join(f"  {i+1}. {c}" for i, c in enumerate(model_changes)) or "  (none)"

    prompt = (
        f"{_CHANGES_JUDGE_SYSTEM}\n\n"
        f"Human changes (ground truth):\n{human_list}\n\n"
        f"Model changes:\n{model_list}\n"
    )

    for attempt in range(retries + 1):
        try:
            client = _get_llm_client()
            response = client.models.generate_content(
                model=model_id,
                contents=[types.Part.from_text(text=prompt)],
                config=types.GenerateContentConfig(
                    response_mime_type="application/json",
                    response_schema=ChangesOutput,
                ),
            )
            result = ChangesOutput.model_validate_json(response.text)
            return {
                "score":       round(result.score, 4),
                "matched":     result.matched,
                "total_human": result.total_human,
                "reason":      result.reason,
                "method":      "llm",
            }
        except Exception as exc:
            logger.warning(
                "LLM changes judge attempt %d/%d failed: %s", attempt + 1, retries + 1, exc
            )
            if attempt < retries:
                time.sleep(2)

    # Fallback: simple count ratio
    ratio = min(len(model_changes), len(human_changes)) / len(human_changes)
    return {
        "score":       round(ratio, 4),
        "matched":     min(len(model_changes), len(human_changes)),
        "total_human": len(human_changes),
        "reason":      "LLM unavailable — fell back to count ratio",
        "method":      "count_fallback",
    }


# ---------------------------------------------------------------------------
# Sheet readers
# ---------------------------------------------------------------------------

def _read_label_value_sheet(ws) -> dict[str, str | None]:
    """Read col A (label) -> col B (value) mapping from a sheet."""
    data: dict[str, str | None] = {}
    for row in ws.iter_rows(values_only=True):
        label = row[0]
        value = row[1] if len(row) > 1 else None
        if label and isinstance(label, str):
            data[label.strip()] = str(value).strip() if value is not None else None
    return data


def _read_changes_sheet(ws) -> list[str]:
    """Read change descriptions from Two Drawings Comparison sheet."""
    changes: list[str] = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == "Change ID":
            header_found = True
            continue
        if not header_found:
            continue
        if row[0] is not None and row[1] is not None:
            changes.append(str(row[1]).strip())
    return changes


# ---------------------------------------------------------------------------
# Per-test validation
# ---------------------------------------------------------------------------

def validate_test(
    n: int,
    use_llm: bool = True,
    judge_model: str = "gemini-2.5-flash",
) -> dict[str, Any] | None:
    """Run validation for a single test. Returns report dict or None if skipped."""
    import openpyxl

    # Paths
    filled_path = RESULTS_DIR / f"test_{n}" / "validation_filled.xlsx"

    template_candidates = [
        VALIDATION_DIR / f"Drawing Data Extraction - Number_ {n}.xlsx",
        VALIDATION_DIR / f"Drawing Data Extraction - Number_ {n}_.xlsx",
    ]
    human_path = next((p for p in template_candidates if p.exists()), None)

    if not filled_path.exists():
        logger.warning("[test %d] validation_filled.xlsx not found — skipping.", n)
        return None
    if human_path is None:
        logger.warning("[test %d] human template not found — skipping.", n)
        return None

    logger.info("[test %d] Validating...", n)

    wb_model  = openpyxl.load_workbook(filled_path)
    wb_human  = openpyxl.load_workbook(human_path)

    model_single = _read_label_value_sheet(wb_model["Single Drawing Data Extraction"])
    human_single = _read_label_value_sheet(wb_human["Single Drawing Data Extraction"])

    model_changes = _read_changes_sheet(wb_model["Two Drawings Comparison"])
    human_changes = _read_changes_sheet(wb_human["Two Drawings Comparison"])

    field_results: dict[str, dict] = {}

    # ── Textual fields (LLM judge) ─────────────────────────────────────────
    for field in TEXTUAL_FIELDS:
        human_val = human_single.get(field)
        model_val = model_single.get(field)

        # Skip if human didn't fill it
        if human_val is None:
            continue

        if use_llm:
            result = llm_judge(field, human_val, model_val, model_id=judge_model)
        else:
            sc = exact_score(human_val, model_val)
            result = {"score": sc, "reason": "exact match (LLM disabled)", "method": "exact"}

        result.update({"human": human_val, "model": model_val, "category": "textual"})
        field_results[field] = result
        logger.info(
            "[test %d] %s: %.2f (%s)",
            n, field, result["score"], result["method"],
        )

    # ── Exact fields ──────────────────────────────────────────────────────
    for field in EXACT_FIELDS:
        human_val = human_single.get(field)
        model_val = model_single.get(field)
        if human_val is None:
            continue
        sc = exact_score(human_val, model_val)
        field_results[field] = {
            "score":    sc,
            "reason":   "exact string match",
            "method":   "exact",
            "human":    human_val,
            "model":    model_val,
            "category": "exact",
        }

    # ── Metric fields (MAPE) ──────────────────────────────────────────────
    metric_scores: list[float] = []
    for field in METRIC_FIELDS:
        human_val = human_single.get(field)
        model_val = model_single.get(field)

        # Skip if human hasn't filled this metric yet
        if human_val is None or human_val in ("None", ""):
            continue

        try:
            h = float(str(human_val).replace(",", "."))
            m = float(str(model_val).replace(",", ".")) if model_val not in (None, "None", "") else 0.0
        except (ValueError, TypeError):
            logger.warning("[test %d] Could not parse metric %r: human=%r model=%r",
                           n, field, human_val, model_val)
            continue

        sc = mape_score(h, m)
        metric_scores.append(sc)
        field_results[field] = {
            "score":    round(sc, 4),
            "human":    h,
            "model":    m,
            "method":   "mape",
            "category": "metric",
            "reason":   f"MAPE score: |{m}-{h}|/{h if h else 1:.4f}",
        }
        logger.info("[test %d] %s: %.2f (MAPE, human=%g model=%g)", n, field, sc, h, m)

    # ── Changes (LLM judge) ───────────────────────────────────────────────
    if use_llm:
        changes_result = llm_judge_changes(human_changes, model_changes, model_id=judge_model)
    else:
        ratio = (
            min(len(model_changes), len(human_changes)) / len(human_changes)
            if human_changes else 1.0
        )
        changes_result = {
            "score":       round(ratio, 4),
            "matched":     min(len(model_changes), len(human_changes)),
            "total_human": len(human_changes),
            "reason":      "count ratio (LLM disabled)",
            "method":      "count",
        }
    changes_result["human_count"] = len(human_changes)
    changes_result["model_count"] = len(model_changes)
    logger.info("[test %d] Changes: %.2f (%d human / %d model)",
                n, changes_result["score"], len(human_changes), len(model_changes))

    # ── Composite score ───────────────────────────────────────────────────
    drawing_block_scores = [
        v["score"] for v in field_results.values()
        if v.get("category") in ("textual", "exact")
    ]
    metric_scores_all = [
        v["score"] for v in field_results.values()
        if v.get("category") == "metric"
    ]

    avg_drawing = (sum(drawing_block_scores) / len(drawing_block_scores)
                   if drawing_block_scores else None)
    avg_metrics = (sum(metric_scores_all) / len(metric_scores_all)
                   if metric_scores_all else None)
    avg_changes = changes_result["score"]

    # Only include components that have data
    components: list[tuple[float, float]] = []
    if avg_drawing is not None:
        components.append((avg_drawing, WEIGHT_DRAWING_BLOCK))
    if avg_metrics is not None:
        components.append((avg_metrics, WEIGHT_METRICS))
    components.append((avg_changes, WEIGHT_CHANGES))

    total_weight = sum(w for _, w in components)
    score_total = sum(s * w for s, w in components) / total_weight if total_weight else 0.0

    report = {
        "test":         n,
        "score_total":  round(score_total, 4),
        "breakdown": {
            "drawing_block": round(avg_drawing, 4) if avg_drawing is not None else None,
            "metrics":       round(avg_metrics, 4) if avg_metrics is not None else None,
            "changes":       round(avg_changes, 4),
        },
        "fields":        field_results,
        "changes":       changes_result,
        "human_changes": human_changes,
        "model_changes": model_changes,
    }

    # Save per-test report
    report_path = RESULTS_DIR / f"test_{n}" / "validation_report.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    logger.info("[test %d] Report saved -> %s (score=%.4f)", n, report_path, score_total)

    return report


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%H:%M:%S",
    )

    parser = argparse.ArgumentParser(
        description="Validate model outputs against human answers (MAPE + LLM-as-a-Judge)"
    )
    parser.add_argument(
        "--tests", type=int, nargs="+",
        default=list(range(41, 51)),
        help="Test numbers to validate (default: all 41-50)",
    )
    parser.add_argument(
        "--no-llm", action="store_true",
        help="Skip LLM judge — use exact match and count ratio only",
    )
    parser.add_argument(
        "--judge-model", default="gemini-2.5-flash",
        help="Gemini model ID for LLM-as-a-Judge (default: gemini-2.5-flash)",
    )
    args = parser.parse_args()

    use_llm = not args.no_llm
    if not use_llm:
        logger.info("LLM judge disabled — using exact match and MAPE only.")

    summary: list[dict] = []
    skipped: list[int]  = []

    for n in args.tests:
        report = validate_test(n, use_llm=use_llm, judge_model=args.judge_model)
        if report is None:
            skipped.append(n)
        else:
            summary.append({
                "test":        report["test"],
                "score_total": report["score_total"],
                "breakdown":   report["breakdown"],
            })

    # Overall summary
    if summary:
        overall = sum(r["score_total"] for r in summary) / len(summary)
        summary_doc = {
            "overall_score": round(overall, 4),
            "tests_evaluated": len(summary),
            "tests_skipped":   skipped,
            "results":         sorted(summary, key=lambda r: r["test"]),
        }
        summary_path = RESULTS_DIR / "validation_summary.json"
        with open(summary_path, "w", encoding="utf-8") as f:
            json.dump(summary_doc, f, indent=2, ensure_ascii=False)

        print(f"\n{'='*50}")
        print(f"VALIDATION COMPLETE")
        print(f"{'='*50}")
        print(f"Overall score: {overall:.4f} ({overall*100:.1f}%)")
        print(f"Tests evaluated: {len(summary)} / {len(args.tests)}")
        if skipped:
            print(f"Skipped (no results yet): {skipped}")
        print(f"\nPer-test scores:")
        for r in sorted(summary, key=lambda x: x["test"]):
            bd = r["breakdown"]
            print(
                f"  Test {r['test']:2d}: {r['score_total']:.4f}"
                f"  (drawing={bd['drawing_block'] or 'N/A':.4f}"  # type: ignore[str-format]
                f"  metrics={bd['metrics'] or 'N/A'}"
                f"  changes={bd['changes']:.4f})"
            )
        print(f"\nSummary saved -> {summary_path}")
    else:
        print("No tests were evaluated. Run scripts/run_batch.py and scripts/generate_validation.py first.")


if __name__ == "__main__":
    main()
