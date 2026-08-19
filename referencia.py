"""
referencia.py — Teste de mapeamento de normas TSS para desenhos CAD Nidec.

Objetivo: Validar se o LLM consegue identificar quais normas TSS se aplicam
a um desenho CAD, usando raciocínio dedutivo de engenharia (não busca literal).

Dois modos de teste:
  1. TEXTO   → extrai texto vetorizado do PDF e envia ao Gemini
  2. FALLBACK → texto + imagem renderizada (ativado quando confidence < threshold)

Uso:
    python referencia.py CADS/13358002_REV_7_draw_1.pdf
    python referencia.py CADS/13358002_REV_7_draw_1.pdf --mode text
    python referencia.py CADS/13358002_REV_7_draw_1.pdf --mode image
    python referencia.py CADS/13358002_REV_7_draw_1.pdf --mode both
    python referencia.py CADS/13358002_REV_7_draw_1.pdf --normas normas.xlsx

Modos disponíveis:
    text   → apenas texto vetorizado (Teste 1)
    image  → texto + imagem renderizada (Teste 2)
    both   → executa os dois e compara os resultados (padrão)
"""

from __future__ import annotations

import argparse
import base64
import json
import logging
import os
import time
from dataclasses import dataclass, field
from pathlib import Path

import fitz
from dotenv import load_dotenv
from google import genai
from google.genai import types
from pydantic import BaseModel, Field

from src.utils.helper_func import pdf_to_images_base64

# ---------------------------------------------------------------------------
# Setup
# ---------------------------------------------------------------------------
load_dotenv()
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

GCP_PROJECT = os.getenv("GCP_PROJECT_ID", "acim-global-data-lake-sandbox")
GCP_REGION  = os.getenv("GCP_REGION", "us-east5")
MODEL_ID    = "gemini-2.5-flash"

# Confidence mínima para considerar o resultado do modo texto como suficiente.
# Abaixo disso, o fallback texto+imagem é acionado automaticamente.
CONFIDENCE_THRESHOLD = 0.70

# Caminho padrão do xlsx. Pode ser sobrescrito via --normas na CLI.
DEFAULT_NORMAS_PATH = Path("normas.xlsx")


# ---------------------------------------------------------------------------
# Carregamento da tabela de normas — fonte: normas.xlsx
# ---------------------------------------------------------------------------

def load_standards_table(xlsx_path: Path = DEFAULT_NORMAS_PATH) -> list[dict]:
    """
    Lê a tabela de normas TSS do arquivo Excel e retorna uma lista de dicts.

    Estrutura esperada da planilha (aba "Notes"):
        Linha 1: vazia
        Linha 2: cabeçalho  (Standard | Content | Category | Compressor Series | Applicability)
        Linha 3+: dados

    Colunas por índice (row[:6]):
        0 → vazia  |  1 → Standard  |  2 → Content  |  3 → Category
        4 → Compressor Series  |  5 → Applicability

    FASE 2 — Google Sheets (substituir este método):
        from googleapiclient.discovery import build
        from google.oauth2 import service_account

        creds = service_account.Credentials.from_service_account_file(
            "service_account.json",
            scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
        )
        service = build("sheets", "v4", credentials=creds)
        result = service.spreadsheets().values().get(
            spreadsheetId=SPREADSHEET_ID,
            range="Notes!B2:F",
        ).execute()
        rows = result.get("values", [])
        return [
            {
                "standard": r[0], "content": r[1], "category": r[2],
                "compressor_series": r[3], "applicability": r[4],
            }
            for r in rows[1:] if len(r) >= 5 and r[0]
        ]
    """
    import openpyxl

    if not xlsx_path.exists():
        raise FileNotFoundError(
            f"Arquivo de normas não encontrado: {xlsx_path}\n"
            "Verifique o caminho ou use --normas para especificar outro arquivo."
        )

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Notes"] if "Notes" in wb.sheetnames else wb.active

    standards = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        _, standard, content, category, compressor_series, applicability = row[:6]
        if not standard:
            continue
        standards.append({
            "standard": str(standard).strip(),
            "content": str(content).strip() if content else "",
            "category": str(category).strip() if category else "",
            "compressor_series": str(compressor_series).strip() if compressor_series else "All",
            "applicability": str(applicability).strip() if applicability else "",
        })

    wb.close()
    logger.info("Tabela de normas carregada: %d entradas de '%s'", len(standards), xlsx_path)
    return standards


# ---------------------------------------------------------------------------
# Pydantic models — structured output
# ---------------------------------------------------------------------------

class AppliedStandard(BaseModel):
    """Uma norma TSS identificada como aplicável ao desenho."""

    standard: str = Field(description="Código da norma, ex: TSS 002369")
    applicability_match: str = Field(
        description="Trecho da coluna Applicability que justifica a aplicação"
    )
    reasoning: str = Field(
        description=(
            "Raciocínio de engenharia de duas etapas: "
            "(1) natureza física do componente, "
            "(2) por que essa norma se aplica"
        )
    )
    confidence: float = Field(description="Confiança de 0.0 a 1.0 desta associação específica")


class StandardsMappingOutput(BaseModel):
    """Saída estruturada do mapeamento de normas para um desenho CAD."""

    component_title: str = Field(
        description="Título extraído do campo TITLE, DOCUMENT TYPE do carimbo"
    )
    component_inference: str = Field(
        description=(
            "Inferência de composição: natureza física, material e função "
            "do componente no ecossistema do compressor"
        )
    )
    applied_standards: list[AppliedStandard] = Field(
        description="Lista de normas TSS identificadas como aplicáveis, em ordem de relevância"
    )
    explicitly_excluded: list[str] = Field(
        description="Normas que foram consideradas e explicitamente descartadas, com breve justificativa"
    )
    overall_confidence: float = Field(
        description="Confiança geral de 0.0 a 1.0 no mapeamento completo"
    )
    needs_visual_confirmation: bool = Field(
        description="True se o texto foi insuficiente e a análise visual do desenho seria necessária"
    )


# ---------------------------------------------------------------------------
# Prompts
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = """\
Você é um especialista sênior em normas técnicas de engenharia para compressores herméticos Embraco/Nidec.

Sua tarefa é identificar quais normas TSS da tabela fornecida se aplicam ao desenho técnico CAD analisado.

## Regra de Associação (OBRIGATÓRIO — duas etapas)

**Etapa 1 — Inferência de Composição:**
Analise o título (TITLE, DOCUMENT TYPE) e o conteúdo do desenho para determinar:
- Qual é a natureza física do componente? (peça usinada, estampada, fundida, montagem, etc.)
- De que material é feito? (aço, alumínio, ferro fundido, cobre, sinterizado, etc.)
- Qual é sua função dentro do compressor? (peça isolada, subconjunto, produto final montado, etc.)

**Etapa 2 — Mapeamento por Categoria de Pertencimento:**
Com base na inferência acima, associe o componente à categoria correta na coluna Applicability.
NÃO faça correspondência literal de palavras — aplique raciocínio dedutivo.

## Exemplo (Golden Rule)
- Título: "Stator - Stack"
- Inferência: componente metálico de aço que forma o motor elétrico, é uma peça individual, não é o compressor montado
- Norma aplicada: TSS 002369 (All metallic components) ✓
- Norma ignorada: TSS 001266 (Compressor assembly) — é uma peça, não uma montagem ✗

## Normas com Applicability "All"
TSS 002470 e TSS 002513 se aplicam a TODOS os desenhos técnicos, sem exceção.
TSS 002420 se aplica a TODOS os componentes (materiais, produtos acabados, embalagens).

## Formato de saída
Retorne JSON estruturado conforme o schema definido.
"""


def _build_standards_context(standards: list[dict]) -> str:
    """Serializa a tabela de normas como bloco de texto para o prompt."""
    lines = ["## Tabela de Normas TSS Disponíveis\n"]
    lines.append(f"{'Standard':<14} {'Category':<22} {'Compressor Series':<30} {'Applicability'}")
    lines.append("-" * 100)
    for s in standards:
        series = s.get("compressor_series", "All")
        lines.append(
            f"{s['standard']:<14} {s['category']:<22} {series:<30} {s['applicability']}"
        )
    return "\n".join(lines)


def _build_text_prompt(pdf_text: str, standards: list[dict]) -> str:
    standards_block = _build_standards_context(standards)
    return (
        f"{SYSTEM_PROMPT}\n\n"
        f"{standards_block}\n\n"
        "## Texto vetorizado do desenho CAD\n\n"
        f"{pdf_text}\n\n"
        "Com base no texto acima, identifique o componente e mapeie todas as normas aplicáveis.\n"
        "Preencha 'needs_visual_confirmation = true' se o texto for insuficiente para ter certeza."
    )


def _build_image_prompt(pdf_text: str, standards: list[dict]) -> str:
    standards_block = _build_standards_context(standards)
    return (
        f"{SYSTEM_PROMPT}\n\n"
        f"{standards_block}\n\n"
        "## Texto vetorizado do desenho CAD (complementar à imagem)\n\n"
        f"{pdf_text}\n\n"
        "A imagem acima mostra o desenho técnico completo.\n"
        "Use tanto o texto quanto a imagem para identificar o componente e mapear todas as normas aplicáveis."
    )


# ---------------------------------------------------------------------------
# LLM client — singleton
# ---------------------------------------------------------------------------

_client: genai.Client | None = None


def _get_client() -> genai.Client:
    """Retorna um cliente Gemini singleton — criado uma única vez e reutilizado."""
    global _client
    if _client is None:
        logger.info(
            "Inicializando cliente Gemini (projeto=%s, região=%s)...",
            GCP_PROJECT, GCP_REGION,
        )
        _client = genai.Client(vertexai=True, project=GCP_PROJECT, location=GCP_REGION)
        logger.info("Cliente Gemini inicializado.")
    return _client


# ---------------------------------------------------------------------------
# LLM calls
# ---------------------------------------------------------------------------

def call_text_only(
    pdf_text: str,
    standards: list[dict],
    model: str = MODEL_ID,
) -> tuple[StandardsMappingOutput, float]:
    """Chama o LLM usando apenas o texto vetorizado do PDF."""
    logger.info("[MODO TEXTO] Enviando texto vetorizado para %s...", model)
    t0 = time.time()

    prompt = _build_text_prompt(pdf_text, standards)
    response = _get_client().models.generate_content(
        model=model,
        contents=[types.Part.from_text(text=prompt)],
        config=types.GenerateContentConfig(
            response_mime_type="application/json",
            response_schema=StandardsMappingOutput,
        ),
    )

    result = StandardsMappingOutput.model_validate_json(response.text)
    latency = time.time() - t0
    logger.info(
        "[MODO TEXTO] Concluído em %.1fs | confidence=%.2f",
        latency, result.overall_confidence,
    )
    return result, latency


def call_text_and_image(
    pdf_text: str,
    pdf_bytes: bytes,
    standards: list[dict],
    model: str = MODEL_ID,
    dpi: int = 150,
) -> tuple[StandardsMappingOutput, float]:
    """Chama o LLM com texto vetorizado + imagem renderizada do PDF."""
    logger.info("[MODO IMAGEM] Renderizando PDF e enviando para %s...", model)
    t0 = time.time()

    images_b64 = pdf_to_images_base64(pdf_bytes, dpi=dpi)
    prompt = _build_image_prompt(pdf_text, standards)

    parts: list[types.Part] = []
    for i, img_b64 in enumerate(images_b64):
        parts.append(
            types.Part.from_bytes(data=base64.b64decode(img_b64), mime_type="image/png")
        )
        parts.append(types.Part.from_text(text=f"[Página {i + 1} do desenho acima]"))
    parts.append(types.Part.from_text(text=prompt))

    response = _get_client().models.generate_content(
        model=model,
        contents=parts,
        config=types.GenerateContentConfig(
            response_mime_type="application/json",
            response_schema=StandardsMappingOutput,
        ),
    )

    result = StandardsMappingOutput.model_validate_json(response.text)
    latency = time.time() - t0
    logger.info(
        "[MODO IMAGEM] Concluído em %.1fs | confidence=%.2f",
        latency, result.overall_confidence,
    )
    return result, latency


# ---------------------------------------------------------------------------
# Extração de texto do PDF
# ---------------------------------------------------------------------------

def extract_text_from_pdf(pdf_bytes: bytes) -> str:
    """Extrai texto vetorizado de todas as páginas do PDF."""
    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
        pages = [page.get_text() for page in doc]
    text = "\n\n".join(pages)
    logger.info("Texto extraído: %d caracteres em %d página(s)", len(text), len(pages))
    return text


# ---------------------------------------------------------------------------
# Renderização dos resultados
# ---------------------------------------------------------------------------

def _separator(char: str = "─", width: int = 72) -> str:
    return char * width


def print_result(label: str, result: StandardsMappingOutput, latency: float) -> None:
    print(f"\n{'═' * 72}")
    print(f"  RESULTADO — {label}")
    print(f"{'═' * 72}")
    print(f"  Componente : {result.component_title}")
    print(f"  Confiança  : {result.overall_confidence:.0%}")
    print(f"  Latência   : {latency:.1f}s")
    print(f"  Confirmação visual necessária: {'SIM' if result.needs_visual_confirmation else 'NÃO'}")
    print()
    print("  INFERÊNCIA DE COMPOSIÇÃO:")
    print(f"  {result.component_inference}")
    print()

    if result.applied_standards:
        print(f"  NORMAS APLICÁVEIS ({len(result.applied_standards)}):")
        print(_separator())
        for s in result.applied_standards:
            print(f"  ✓ {s.standard}  [{s.confidence:.0%}]")
            print(f"    Aplicabilidade : {s.applicability_match}")
            print(f"    Raciocínio     : {s.reasoning}")
            print()
    else:
        print("  Nenhuma norma identificada.")

    if result.explicitly_excluded:
        print(_separator())
        print("  NORMAS DESCARTADAS:")
        for item in result.explicitly_excluded:
            print(f"  ✗ {item}")

    print(_separator("═"))


def save_json_result(output_path: Path, data: dict) -> None:
    output_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    logger.info("Resultado salvo em: %s", output_path)


# ---------------------------------------------------------------------------
# Orquestrador principal
# ---------------------------------------------------------------------------

@dataclass
class TestSession:
    pdf_path: Path
    mode: str                            # "text" | "image" | "both"
    confidence_threshold: float = CONFIDENCE_THRESHOLD
    model: str = MODEL_ID
    dpi: int = 150
    normas_path: Path = field(default_factory=lambda: DEFAULT_NORMAS_PATH)

    # preenchidos durante a execução
    pdf_bytes: bytes = field(default=b"", repr=False)
    pdf_text: str = field(default="")
    standards: list[dict] = field(default_factory=list)


def run(session: TestSession) -> dict:
    """Executa o teste conforme o modo solicitado. Retorna dict com todos os resultados."""
    logger.info("Carregando PDF: %s", session.pdf_path)
    session.pdf_bytes = session.pdf_path.read_bytes()
    session.pdf_text  = extract_text_from_pdf(session.pdf_bytes)
    session.standards = load_standards_table(session.normas_path)

    output: dict = {
        "pdf": session.pdf_path.name,
        "model": session.model,
        "mode": session.mode,
        "normas_source": str(session.normas_path),
        "confidence_threshold": session.confidence_threshold,
    }

    # ── Modo texto ──────────────────────────────────────────────────────────
    if session.mode in ("text", "both"):
        text_result, text_latency = call_text_only(
            session.pdf_text, session.standards, model=session.model
        )
        print_result("TEXTO VETORIZADO", text_result, text_latency)
        output["text_result"] = {
            "latency_s": round(text_latency, 2),
            **json.loads(text_result.model_dump_json()),
        }

        # Lógica de fallback automático no modo "both"
        if session.mode == "both":
            low_confidence = text_result.overall_confidence < session.confidence_threshold
            needs_visual   = text_result.needs_visual_confirmation

            if low_confidence or needs_visual:
                reason = []
                if low_confidence:
                    reason.append(
                        f"confidence={text_result.overall_confidence:.2f} "
                        f"< threshold={session.confidence_threshold:.2f}"
                    )
                if needs_visual:
                    reason.append("LLM sinalizou necessidade de confirmação visual")
                logger.warning("FALLBACK acionado: %s", " | ".join(reason))
                output["fallback_triggered"] = True
                output["fallback_reason"] = reason

                image_result, image_latency = call_text_and_image(
                    session.pdf_text, session.pdf_bytes, session.standards,
                    model=session.model, dpi=session.dpi,
                )
                print_result("TEXTO + IMAGEM (fallback)", image_result, image_latency)
                output["image_result"] = {
                    "latency_s": round(image_latency, 2),
                    **json.loads(image_result.model_dump_json()),
                }
            else:
                logger.info(
                    "Texto suficiente (confidence=%.2f >= %.2f). Fallback NÃO acionado.",
                    text_result.overall_confidence,
                    session.confidence_threshold,
                )
                output["fallback_triggered"] = False

    # ── Modo imagem direto ───────────────────────────────────────────────────
    elif session.mode == "image":
        image_result, image_latency = call_text_and_image(
            session.pdf_text, session.pdf_bytes, session.standards,
            model=session.model, dpi=session.dpi,
        )
        print_result("TEXTO + IMAGEM", image_result, image_latency)
        output["image_result"] = {
            "latency_s": round(image_latency, 2),
            **json.loads(image_result.model_dump_json()),
        }

    return output


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Teste de mapeamento de normas TSS para desenhos CAD Nidec."
    )
    parser.add_argument(
        "pdf",
        type=Path,
        help="Caminho para o PDF do desenho CAD",
    )
    parser.add_argument(
        "--mode",
        choices=["text", "image", "both"],
        default="both",
        help=(
            "text  → só texto vetorizado | "
            "image → texto + imagem | "
            "both  → texto primeiro, fallback para imagem se necessário (padrão)"
        ),
    )
    parser.add_argument(
        "--normas",
        type=Path,
        default=DEFAULT_NORMAS_PATH,
        help=f"Caminho para o arquivo normas.xlsx (padrão: {DEFAULT_NORMAS_PATH})",
    )
    parser.add_argument(
        "--model",
        default=MODEL_ID,
        help=f"Modelo Gemini a usar (padrão: {MODEL_ID})",
    )
    parser.add_argument(
        "--threshold",
        type=float,
        default=CONFIDENCE_THRESHOLD,
        help=f"Confidence mínima para evitar fallback (padrão: {CONFIDENCE_THRESHOLD})",
    )
    parser.add_argument(
        "--dpi",
        type=int,
        default=150,
        help="DPI para renderização do PDF como imagem (padrão: 150)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Caminho para salvar o resultado em JSON (opcional)",
    )
    args = parser.parse_args()

    if not args.pdf.exists():
        parser.error(f"Arquivo não encontrado: {args.pdf}")

    session = TestSession(
        pdf_path=args.pdf,
        mode=args.mode,
        confidence_threshold=args.threshold,
        model=args.model,
        dpi=args.dpi,
        normas_path=args.normas,
    )

    output = run(session)

    if args.output:
        save_json_result(args.output, output)
    else:
        default_out = args.pdf.with_name(args.pdf.stem + "_normas_result.json")
        save_json_result(default_out, output)


if __name__ == "__main__":
    main()
